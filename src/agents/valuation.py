"""Valuation agent — DCF + Comps + Sensitivity + Football Field.

Takes a ModelOutputs from the model agent and runs the full valuation pipeline
via the deterministic calculator. Writes a Valuation tab into the Excel model
and returns the FixedNumbers object.
"""
from __future__ import annotations
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..engine.calculator import (
    FixedNumbers,
    build_fixed_numbers,
    compute_dcf_price,
)
from ..agents.model import ModelOutputs


# Reuse style constants
BOLD = Font(name="Arial", size=10, bold=True, color="000000")
HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=14, bold=True, color="000000")
BLUE = Font(name="Arial", size=10, color="0000FF")
DARKBLUE = PatternFill("solid", start_color="1F4E78")
LIGHTBLUE = PatternFill("solid", start_color="DDEBF7")
GREEN_FILL = PatternFill("solid", start_color="C6EFCE")
YELLOW = PatternFill("solid", start_color="FFFF00")
RIGHT = Alignment(horizontal="right")

FMT_DOLLAR = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_MULT = '0.0"x"'
FMT_PRICE = '$#,##0;($#,##0);"-"'


# Peer multiple defaults — used when no peers explicitly provided.
# Pulled from the broader infrastructure-services sector backdrop discussed in the AGX work.
# In production, the screener provides actual peer-derived medians.
DEFAULT_PEER_MULTIPLES = {
    "median_ev_ebitda": 18.0,
    "p75_ev_ebitda": 25.0,
    "median_pe": 28.0,
}


def value_ticker(
    model: ModelOutputs,
    current_price: float,
    wacc: float = 0.11,
    terminal_growth: float = 0.025,
    exit_multiple: float = 12.0,
    peer_multiples: Optional[dict] = None,
    consensus_range: Optional[tuple[float, float]] = None,
    excel_path: Optional[Path] = None,
    n_peers_resolved: int = 0,
    cohort_outlier: bool = False,
) -> FixedNumbers:
    """Run the full DCF + comps valuation and return FixedNumbers.

    If excel_path supplied, appends a Valuation sheet to that workbook.
    """
    peers = peer_multiples or DEFAULT_PEER_MULTIPLES

    # Terminal year FCF — last UFCF in the stream
    if not model.ufcf_stream:
        raise ValueError("Empty UFCF stream — model did not project")
    terminal_fcf = model.ufcf_stream[-1]

    # Sensitivity ranges — compute corner cases
    # Low: high WACC + low terminal g (Gordon) / low exit multiple
    p_low_g, _, _ = compute_dcf_price(
        ufcf_stream=model.ufcf_stream,
        terminal_year_fcf=terminal_fcf,
        terminal_growth=max(0.01, terminal_growth - 0.01),
        exit_multiple=exit_multiple,
        terminal_year_ebitda=model.terminal_ebitda,
        wacc=wacc + 0.02,
        cash_plus_investments=model.cash_plus_investments,
        debt=model.debt,
        diluted_shares_mm=model.diluted_shares_mm,
    )
    # High: low WACC + high terminal g
    p_high_g, _, _ = compute_dcf_price(
        ufcf_stream=model.ufcf_stream,
        terminal_year_fcf=terminal_fcf,
        terminal_growth=terminal_growth + 0.01,
        exit_multiple=exit_multiple,
        terminal_year_ebitda=model.terminal_ebitda,
        wacc=max(0.05, wacc - 0.02),
        cash_plus_investments=model.cash_plus_investments,
        debt=model.debt,
        diluted_shares_mm=model.diluted_shares_mm,
    )
    # Exit method low/high — vary exit multiple ±25%
    _, p_low_e, _ = compute_dcf_price(
        ufcf_stream=model.ufcf_stream,
        terminal_year_fcf=terminal_fcf,
        terminal_growth=terminal_growth,
        exit_multiple=exit_multiple * 0.75,
        terminal_year_ebitda=model.terminal_ebitda,
        wacc=wacc + 0.01,
        cash_plus_investments=model.cash_plus_investments,
        debt=model.debt,
        diluted_shares_mm=model.diluted_shares_mm,
    )
    _, p_high_e, _ = compute_dcf_price(
        ufcf_stream=model.ufcf_stream,
        terminal_year_fcf=terminal_fcf,
        terminal_growth=terminal_growth,
        exit_multiple=exit_multiple * 1.25,
        terminal_year_ebitda=model.terminal_ebitda,
        wacc=max(0.05, wacc - 0.01),
        cash_plus_investments=model.cash_plus_investments,
        debt=model.debt,
        diluted_shares_mm=model.diluted_shares_mm,
    )

    fixed = build_fixed_numbers(
        ticker=model.last_historical_period[:4] if model.last_historical_period else "UNK",
        current_price=current_price,
        diluted_shares_mm=model.diluted_shares_mm,
        ufcf_stream=model.ufcf_stream,
        terminal_year_fcf=terminal_fcf,
        terminal_year_ebitda=model.terminal_ebitda,
        wacc=wacc,
        terminal_growth=terminal_growth,
        exit_multiple=exit_multiple,
        cash_plus_investments=model.cash_plus_investments,
        debt=model.debt,
        target_ebitda=model.fy1_ebitda,
        target_net_income=model.fy1_net_income,
        peer_median_ev_ebitda=peers["median_ev_ebitda"],
        peer_75th_ev_ebitda=peers["p75_ev_ebitda"],
        peer_median_pe=peers["median_pe"],
        consensus_low=consensus_range[0] if consensus_range else None,
        consensus_high=consensus_range[1] if consensus_range else None,
        dcf_gordon_low=p_low_g,
        dcf_gordon_high=p_high_g,
        dcf_exit_low=p_low_e,
        dcf_exit_high=p_high_e,
        n_peers_resolved=n_peers_resolved,
        cohort_outlier=cohort_outlier,
    )

    if excel_path is not None and excel_path.exists():
        _append_valuation_sheets(excel_path, fixed, model, wacc, terminal_growth, exit_multiple, peers, consensus_range)

    return fixed


def _put(ws, addr, val, font=None, fmt=None, fill=None, align=None):
    cell = ws[addr]
    cell.value = val
    if font: cell.font = font
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    if align: cell.alignment = align


def _append_valuation_sheets(
    excel_path: Path,
    fixed: FixedNumbers,
    model: ModelOutputs,
    wacc: float,
    terminal_growth: float,
    exit_multiple: float,
    peers: dict,
    consensus_range: Optional[tuple[float, float]],
) -> None:
    """Add DCF, Comps, Football Field sheets to the existing workbook."""
    wb = load_workbook(excel_path)

    # ===== DCF tab =====
    if "DCF" in wb.sheetnames:
        del wb["DCF"]
    dcf = wb.create_sheet("DCF")
    dcf.column_dimensions["A"].width = 38
    for c in "BCDEFGH": dcf.column_dimensions[c].width = 14

    _put(dcf, "A1", f"{fixed.ticker} — DCF Valuation", TITLE)
    _put(dcf, "A3", "INPUTS", BOLD, fill=LIGHTBLUE)
    _put(dcf, "A4", "  WACC"); _put(dcf, "B4", wacc, BLUE, FMT_PCT, fill=YELLOW)
    _put(dcf, "A5", "  Terminal growth"); _put(dcf, "B5", terminal_growth, BLUE, FMT_PCT, fill=YELLOW)
    _put(dcf, "A6", "  Exit EV/EBITDA"); _put(dcf, "B6", exit_multiple, BLUE, FMT_MULT, fill=YELLOW)
    _put(dcf, "A7", "  Cash + investments ($mm)"); _put(dcf, "B7", model.cash_plus_investments, BLUE, FMT_DOLLAR, fill=YELLOW)
    _put(dcf, "A8", "  Debt ($mm)"); _put(dcf, "B8", model.debt, BLUE, FMT_DOLLAR, fill=YELLOW)
    _put(dcf, "A9", "  Diluted shares (mm)"); _put(dcf, "B9", model.diluted_shares_mm, BLUE, fill=YELLOW)

    _put(dcf, "A11", "UFCF STREAM ($mm)", BOLD, fill=LIGHTBLUE)
    for i, ufcf in enumerate(model.ufcf_stream):
        _put(dcf, f"A{12+i}", f"  FY+{i+1}")
        _put(dcf, f"B{12+i}", ufcf, fmt=FMT_DOLLAR, align=RIGHT)

    _put(dcf, "A18", "OUTPUTS ($/share)", BOLD, fill=LIGHTBLUE)
    _put(dcf, "A19", "  DCF Gordon Growth"); _put(dcf, "B19", fixed.dcf_price_gordon, BOLD, FMT_PRICE, align=RIGHT)
    _put(dcf, "A20", "  DCF Exit Multiple"); _put(dcf, "B20", fixed.dcf_price_exit, BOLD, FMT_PRICE, align=RIGHT)
    _put(dcf, "A21", "  Blended"); _put(dcf, "B21", fixed.dcf_price_blended, BOLD, FMT_PRICE, fill=GREEN_FILL, align=RIGHT)

    # ===== Comps tab =====
    if "Comps" in wb.sheetnames:
        del wb["Comps"]
    cw = wb.create_sheet("Comps")
    cw.column_dimensions["A"].width = 38
    for c in "BCDEFGH": cw.column_dimensions[c].width = 14

    _put(cw, "A1", f"{fixed.ticker} — Comparable Companies", TITLE)
    _put(cw, "A3", "PEER MULTIPLES (assumed / derived)", BOLD, fill=LIGHTBLUE)
    _put(cw, "A4", "  Median EV/EBITDA"); _put(cw, "B4", peers["median_ev_ebitda"], BLUE, FMT_MULT, fill=YELLOW)
    _put(cw, "A5", "  75th pct EV/EBITDA"); _put(cw, "B5", peers["p75_ev_ebitda"], BLUE, FMT_MULT, fill=YELLOW)
    _put(cw, "A6", "  Median P/E"); _put(cw, "B6", peers["median_pe"], BLUE, FMT_MULT, fill=YELLOW)

    _put(cw, "A8", "TARGET METRICS ($mm)", BOLD, fill=LIGHTBLUE)
    _put(cw, "A9", "  FY+1 EBITDA"); _put(cw, "B9", model.fy1_ebitda, fmt=FMT_DOLLAR, align=RIGHT)
    _put(cw, "A10", "  FY+1 Net income"); _put(cw, "B10", model.fy1_net_income, fmt=FMT_DOLLAR, align=RIGHT)
    _put(cw, "A11", "  FY+1 EPS"); _put(cw, "B11", model.fy1_eps, fmt='$#,##0.00', align=RIGHT)

    _put(cw, "A13", "COMPS-IMPLIED PRICES ($/share)", BOLD, fill=LIGHTBLUE)
    _put(cw, "A14", "  Median EV/EBITDA"); _put(cw, "B14", fixed.comps_price_median_ev_ebitda, BOLD, FMT_PRICE, align=RIGHT)
    _put(cw, "A15", "  75th pct EV/EBITDA"); _put(cw, "B15", fixed.comps_price_75th_ev_ebitda, BOLD, FMT_PRICE, align=RIGHT)
    _put(cw, "A16", "  Median P/E"); _put(cw, "B16", fixed.comps_price_median_pe, BOLD, FMT_PRICE, align=RIGHT)

    # ===== Football Field tab =====
    if "Football_Field" in wb.sheetnames:
        del wb["Football_Field"]
    ff = wb.create_sheet("Football_Field")
    ff.column_dimensions["A"].width = 38
    for c in "BCDEFG": ff.column_dimensions[c].width = 14

    _put(ff, "A1", f"{fixed.ticker} — Valuation Football Field", TITLE)
    _put(ff, "A3", "Method", HEADER, fill=DARKBLUE, align=Alignment(horizontal="center"))
    _put(ff, "B3", "Low", HEADER, fill=DARKBLUE, align=Alignment(horizontal="center"))
    _put(ff, "C3", "Mid", HEADER, fill=DARKBLUE, align=Alignment(horizontal="center"))
    _put(ff, "D3", "High", HEADER, fill=DARKBLUE, align=Alignment(horizontal="center"))
    _put(ff, "E3", "Upside vs $" + f"{fixed.current_price:.0f}", HEADER, fill=DARKBLUE, align=Alignment(horizontal="center"))

    rows = [
        ("DCF Gordon Growth", fixed.dcf_gordon_range),
        ("DCF Exit Multiple", fixed.dcf_exit_range),
        ("Comps (multi-method)", fixed.comps_range),
    ]
    if fixed.consensus_range is not None:
        rows.append(("Sell-side consensus", fixed.consensus_range))

    for i, (name, (lo, mid, hi)) in enumerate(rows):
        r = 4 + i
        _put(ff, f"A{r}", name)
        _put(ff, f"B{r}", lo, fmt=FMT_PRICE, align=RIGHT)
        _put(ff, f"C{r}", mid, BOLD, fmt=FMT_PRICE, align=RIGHT)
        _put(ff, f"D{r}", hi, fmt=FMT_PRICE, align=RIGHT)
        upside = mid / fixed.current_price - 1 if fixed.current_price > 0 else 0
        _put(ff, f"E{r}", upside, fmt=FMT_PCT, align=RIGHT)

    summary_r = 4 + len(rows) + 2
    _put(ff, f"A{summary_r}", "BLENDED 12-MO PRICE TARGET", BOLD, fill=LIGHTBLUE)
    _put(ff, f"B{summary_r}", fixed.price_target_12m, BOLD, FMT_PRICE, fill=GREEN_FILL, align=RIGHT)
    _put(ff, f"A{summary_r+1}", "  Current price")
    _put(ff, f"B{summary_r+1}", fixed.current_price, fmt=FMT_PRICE, align=RIGHT)
    _put(ff, f"A{summary_r+2}", "  Expected return", BOLD)
    _put(ff, f"B{summary_r+2}", fixed.expected_return, BOLD, FMT_PCT, fill=GREEN_FILL, align=RIGHT)
    _put(ff, f"A{summary_r+3}", "  Rating", BOLD)
    _put(ff, f"B{summary_r+3}", fixed.rating, BOLD, fill=GREEN_FILL, align=Alignment(horizontal="center"))

    wb.save(excel_path)


if __name__ == "__main__":
    # Quick smoke test: synthetic ModelOutputs
    fake_model = ModelOutputs(
        fy1_revenue=1185,
        fy1_ebitda=165,
        fy1_net_income=128,
        fy1_eps=9.20,
        fy5_revenue=1654,
        fy5_ebitda=176,
        ufcf_stream=[95, 123, 129, 133, 133],
        terminal_ebitda=176,
        diluted_shares_mm=14.0,
        cash_plus_investments=895,
        debt=0,
        last_historical_period="FY2026",
        notes=[],
    )
    fixed = value_ticker(fake_model, current_price=700.0, peer_multiples={
        "median_ev_ebitda": 14.8, "p75_ev_ebitda": 21.2, "median_pe": 27.9,
    }, consensus_range=(220, 321))
    print(f"Rating: {fixed.rating}")
    print(f"PT 12m: ${fixed.price_target_12m:.0f}")
    print(f"Expected return: {fixed.expected_return*100:+.1f}%")
