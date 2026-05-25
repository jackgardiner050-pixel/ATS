"""Financial model agent — builds an Excel 3-statement projection model.

Generic port of build_agx_model.py — parameterised by ticker rather than
hard-coded for AGX. Produces the inputs the valuation agent needs (UFCF
stream, terminal EBITDA, projected EPS, etc.).
"""
from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
import math

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..data.edgar_client import CompanyData, extract_historical_metrics


# Style constants
BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
GREEN = Font(name="Arial", size=10, color="008000")
BOLD = Font(name="Arial", size=10, bold=True, color="000000")
HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=14, bold=True, color="000000")
YELLOW = PatternFill("solid", start_color="FFFF00")
DARKBLUE = PatternFill("solid", start_color="1F4E78")
LIGHTGREY = PatternFill("solid", start_color="F2F2F2")
LIGHTBLUE = PatternFill("solid", start_color="DDEBF7")
RIGHT = Alignment(horizontal="right")

FMT_DOLLAR = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_EPS = '$#,##0.00;($#,##0.00);"-"'


@dataclass
class ModelOutputs:
    """Key projected metrics the valuation agent needs."""
    fy1_revenue: float
    fy1_ebitda: float
    fy1_net_income: float
    fy1_eps: float
    fy5_revenue: float
    fy5_ebitda: float
    ufcf_stream: list[float]
    terminal_ebitda: float
    diluted_shares_mm: float
    cash_plus_investments: float
    debt: float
    last_historical_period: str
    notes: list[str]


def _years_from_metrics(metrics: dict) -> list[str]:
    """Sort the historical periods in ascending order."""
    return sorted(metrics.keys())


def _safe_get(d: dict, key: str, default: float = 0.0) -> float:
    v = d.get(key)
    return float(v) if v is not None and not math.isnan(float(v)) else default


def project_financials(
    metrics: dict,
    growth_decay: list[float] | None = None,
    gm_normalisation: float = 0.0,        # pp absolute gm shift over 5y
    sga_pct_default: float = 0.18,
    tax_rate_default: float = 0.21,
    capex_pct_rev_default: float = 0.04,
) -> dict:
    """Build 5-year projection from historicals.

    Returns a dict of projected periods → metrics, plus a 'meta' key with
    the assumptions used.
    """
    if growth_decay is None:
        growth_decay = [0.18, 0.14, 0.10, 0.07, 0.05]

    years = _years_from_metrics(metrics)
    if not years:
        return {}

    # Last historical year
    last = years[-1]
    last_m = metrics[last]

    base_revenue = _safe_get(last_m, "revenue")
    if base_revenue <= 0:
        return {}

    # Derive SG&A % and capex % from last historical year; fall back to defaults
    last_sga = _safe_get(last_m, "sga")
    sga_pct = abs(last_sga) / base_revenue if last_sga else sga_pct_default

    last_capex = _safe_get(last_m, "capex")
    capex_pct_rev = abs(last_capex) / base_revenue if last_capex else capex_pct_rev_default

    # Compute historical metrics
    gm_history = []
    for y in years:
        m = metrics[y]
        rev = _safe_get(m, "revenue")
        gp = _safe_get(m, "gross_profit")
        if rev > 0 and gp > 0:
            gm_history.append(gp / rev)
    base_gm = gm_history[-1] if gm_history else 0.25

    # Project
    projections = {}
    prev_rev = base_revenue
    prev_gm = base_gm

    for i, g in enumerate(growth_decay):
        rev = prev_rev * (1 + g)
        gm = prev_gm - (gm_normalisation / 5) if gm_normalisation else prev_gm
        gp = rev * gm
        sga = rev * sga_pct
        ebit = gp - sga
        tax = ebit * tax_rate_default
        nopat = ebit - tax
        # Other income — use last historical other_income flat
        other_income = _safe_get(last_m, "other_income")
        ebt = ebit + other_income
        net_income = ebt * (1 - tax_rate_default)

        # D&A: hold proportional to last historical
        last_dep = _safe_get(last_m, "depreciation")
        last_rev = _safe_get(last_m, "revenue")
        dep = last_dep * (rev / last_rev) if last_rev > 0 else last_dep
        ebitda = ebit + dep
        capex = rev * capex_pct_rev
        # NWC change — simple proxy: 5% of revenue change
        nwc_change = (rev - prev_rev) * 0.05
        ufcf = nopat + dep - capex - nwc_change

        projections[f"FY+{i+1}"] = {
            "revenue": rev,
            "gross_margin": gm,
            "gross_profit": gp,
            "sga": sga,
            "ebit": ebit,
            "ebit_margin": ebit / rev if rev > 0 else 0,
            "tax": tax,
            "nopat": nopat,
            "net_income": net_income,
            "depreciation": dep,
            "ebitda": ebitda,
            "ebitda_margin": ebitda / rev if rev > 0 else 0,
            "capex": capex,
            "nwc_change": nwc_change,
            "ufcf": ufcf,
        }
        prev_rev = rev
        prev_gm = gm

    projections["_meta"] = {
        "growth_decay": growth_decay,
        "base_revenue": base_revenue,
        "base_gross_margin": base_gm,
        "gm_normalisation": gm_normalisation,
        "sga_pct": sga_pct,
        "tax_rate": tax_rate_default,
        "capex_pct_rev": capex_pct_rev,
    }
    return projections


def build_model(
    cd: CompanyData,
    output_path: Path,
    growth_decay: list[float] | None = None,
    gm_normalisation: float = 0.02,
    diluted_shares_mm: float = 0.0,
) -> ModelOutputs:
    """Build the Excel model for a ticker and return key projections."""
    metrics = extract_historical_metrics(cd)
    projections = project_financials(
        metrics=metrics,
        growth_decay=growth_decay,
        gm_normalisation=gm_normalisation,
    )

    notes = list(cd.notes)

    if not projections:
        notes.append("no_projections_built — historical data insufficient")

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"

    ws.column_dimensions["A"].width = 38
    for col in "BCDEFGHIJKLMN":
        ws.column_dimensions[col].width = 14

    ws["A1"] = f"{cd.ticker} — {cd.name}"
    ws["A1"].font = TITLE
    ws["A2"] = f"Financial Model — SIC {cd.sic} ({cd.sic_description})"
    ws["A2"].font = BOLD
    ws["A3"] = f"FYE {cd.fiscal_year_end} | Latest 10-K filed {cd.latest_10k_date}"
    ws["A3"].font = Font(name="Arial", size=9, italic=True)

    # Historical + projected periods
    hist_years = _years_from_metrics(metrics)
    proj_years = [k for k in projections.keys() if not k.startswith("_")]
    all_years = hist_years + proj_years

    # Header row
    row = 5
    ws.cell(row=row, column=1).value = "Period"
    ws.cell(row=row, column=1).font = HEADER
    ws.cell(row=row, column=1).fill = DARKBLUE
    for i, y in enumerate(all_years):
        c = ws.cell(row=row, column=2 + i)
        c.value = y
        c.font = HEADER
        c.fill = DARKBLUE
        c.alignment = Alignment(horizontal="center")

    # Helper to write a line
    def write_line(row: int, label: str, accessor, fmt=FMT_DOLLAR, bold=False, fill=None):
        ws.cell(row=row, column=1).value = label
        if bold:
            ws.cell(row=row, column=1).font = BOLD
        if fill:
            ws.cell(row=row, column=1).fill = fill
        for i, y in enumerate(all_years):
            src = metrics.get(y, {}) if y in hist_years else projections.get(y, {})
            val = accessor(src)
            c = ws.cell(row=row, column=2 + i)
            c.value = val
            c.number_format = fmt
            c.alignment = RIGHT
            if val is None or (isinstance(val, float) and math.isnan(val)):
                c.value = None
            else:
                c.font = BLUE if y in hist_years else BLACK

    row = 7
    ws.cell(row=row, column=1).value = "INCOME STATEMENT"
    ws.cell(row=row, column=1).font = BOLD
    ws.cell(row=row, column=1).fill = LIGHTBLUE
    row += 1

    write_line(row, "  Revenue", lambda s: _safe_get(s, "revenue", None) or s.get("revenue")); row += 1
    write_line(row, "  Cost of revenue", lambda s: -_safe_get(s, "cost_of_revenue") or None); row += 1
    write_line(row, "  Gross profit", lambda s: _safe_get(s, "gross_profit", None) or s.get("gross_profit")); row += 1
    write_line(row, "  Gross margin %", lambda s: (s.get("gross_margin") if "gross_margin" in s else (s.get("gross_profit") or 0) / (s.get("revenue") or 1) if s.get("revenue") else None), fmt=FMT_PCT); row += 1
    write_line(row, "  SG&A", lambda s: -_safe_get(s, "sga") or None); row += 1
    write_line(row, "  Operating income (EBIT)", lambda s: s.get("ebit") or s.get("operating_income")); row += 1
    write_line(row, "  Net income", lambda s: s.get("net_income")); row += 1
    write_line(row, "  Diluted EPS", lambda s: s.get("diluted_eps"), fmt=FMT_EPS); row += 2

    ws.cell(row=row, column=1).value = "CASH FLOW"
    ws.cell(row=row, column=1).font = BOLD
    ws.cell(row=row, column=1).fill = LIGHTBLUE
    row += 1
    write_line(row, "  D&A", lambda s: s.get("depreciation")); row += 1
    write_line(row, "  Capex", lambda s: -_safe_get(s, "capex") or None); row += 1
    write_line(row, "  Operating cash flow", lambda s: s.get("operating_cash_flow")); row += 1
    write_line(row, "  EBITDA", lambda s: s.get("ebitda")); row += 1
    write_line(row, "  Unlevered FCF (proj)", lambda s: s.get("ufcf")); row += 2

    ws.cell(row=row, column=1).value = "BALANCE SHEET (latest)"
    ws.cell(row=row, column=1).font = BOLD
    ws.cell(row=row, column=1).fill = LIGHTBLUE
    row += 1
    write_line(row, "  Cash + investments", lambda s: (s.get("cash") or 0) + (s.get("short_term_investments") or 0) if s.get("cash") else None); row += 1
    write_line(row, "  Total assets", lambda s: s.get("total_assets")); row += 1
    write_line(row, "  Debt", lambda s: s.get("debt")); row += 1
    write_line(row, "  Equity", lambda s: s.get("total_equity")); row += 2

    # Assumption block
    ws.cell(row=row, column=1).value = "PROJECTION ASSUMPTIONS"
    ws.cell(row=row, column=1).font = BOLD
    ws.cell(row=row, column=1).fill = LIGHTBLUE
    row += 1
    meta = projections.get("_meta", {})
    ws.cell(row=row, column=1).value = "  Growth decay (FY+1 to FY+5)"
    for i, g in enumerate(meta.get("growth_decay", [])):
        c = ws.cell(row=row, column=2 + i)
        c.value = g
        c.font = BLUE
        c.number_format = FMT_PCT
        c.fill = YELLOW
    row += 1
    ws.cell(row=row, column=1).value = "  SG&A % of revenue"
    c = ws.cell(row=row, column=2); c.value = meta.get("sga_pct"); c.number_format = FMT_PCT; c.font = BLUE; c.fill = YELLOW
    row += 1
    ws.cell(row=row, column=1).value = "  Tax rate"
    c = ws.cell(row=row, column=2); c.value = meta.get("tax_rate"); c.number_format = FMT_PCT; c.font = BLUE; c.fill = YELLOW
    row += 1
    ws.cell(row=row, column=1).value = "  Capex % of revenue"
    c = ws.cell(row=row, column=2); c.value = meta.get("capex_pct_rev"); c.number_format = FMT_PCT; c.font = BLUE; c.fill = YELLOW

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # Compose ModelOutputs
    last_hist = metrics.get(hist_years[-1]) if hist_years else {}
    last_proj = projections.get(proj_years[-1]) if proj_years else {}
    first_proj = projections.get(proj_years[0]) if proj_years else {}

    ufcf_stream = []
    terminal_ebitda = 0.0
    fy1_revenue = 0.0
    fy1_ebitda = 0.0
    fy1_ni = 0.0
    fy5_revenue = 0.0
    fy5_ebitda = 0.0

    if proj_years:
        ufcf_stream = [projections[y]["ufcf"] for y in proj_years]
        terminal_ebitda = last_proj.get("ebitda", 0.0)
        fy1_revenue = first_proj.get("revenue", 0.0)
        fy1_ebitda = first_proj.get("ebitda", 0.0)
        fy1_ni = first_proj.get("net_income", 0.0)
        fy5_revenue = last_proj.get("revenue", 0.0)
        fy5_ebitda = last_proj.get("ebitda", 0.0)

    cash_plus_inv = (_safe_get(last_hist, "cash") + _safe_get(last_hist, "short_term_investments")) / 1e6 if last_hist else 0.0
    debt = _safe_get(last_hist, "debt") / 1e6 if last_hist else 0.0

    # Convert from $ to $mm (EDGAR data always in $; projections inherit that scale)
    # Threshold >1e6 covers all real companies (sub-$1M revenue not in universe)
    if fy1_revenue > 1e6:
        fy1_revenue /= 1e6
        fy1_ebitda /= 1e6
        fy1_ni /= 1e6
        fy5_revenue /= 1e6
        fy5_ebitda /= 1e6
        terminal_ebitda /= 1e6
        ufcf_stream = [u / 1e6 for u in ufcf_stream]

    fy1_eps = fy1_ni / diluted_shares_mm if diluted_shares_mm > 0 else 0.0

    return ModelOutputs(
        fy1_revenue=fy1_revenue,
        fy1_ebitda=fy1_ebitda,
        fy1_net_income=fy1_ni,
        fy1_eps=fy1_eps,
        fy5_revenue=fy5_revenue,
        fy5_ebitda=fy5_ebitda,
        ufcf_stream=ufcf_stream,
        terminal_ebitda=terminal_ebitda,
        diluted_shares_mm=diluted_shares_mm,
        cash_plus_investments=cash_plus_inv,
        debt=debt,
        last_historical_period=hist_years[-1] if hist_years else "",
        notes=notes,
    )


if __name__ == "__main__":
    # Local-only smoke test
    import sys
    from ..data.edgar_client import fetch_company
    ticker = sys.argv[1] if len(sys.argv) > 1 else "AGX"
    cd = fetch_company(ticker)
    out_path = Path(f"runs/{ticker}/model.xlsx")
    outputs = build_model(cd, out_path, diluted_shares_mm=14.0)
    print(f"Built model: {out_path}")
    print(f"FY+1 revenue: ${outputs.fy1_revenue:,.0f}M")
    print(f"FY+1 EBITDA:  ${outputs.fy1_ebitda:,.0f}M")
    print(f"FY+1 NI:      ${outputs.fy1_net_income:,.0f}M")
    print(f"UFCF stream:  {[round(u, 1) for u in outputs.ufcf_stream]}")
    print(f"Terminal EBITDA: ${outputs.terminal_ebitda:,.0f}M")
    print(f"Cash + inv:   ${outputs.cash_plus_investments:,.0f}M")
    print(f"Debt:         ${outputs.debt:,.0f}M")
